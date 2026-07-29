"""
API роутер для отримання списку всіх документів (об'єднаний список).

Ендпоінти:
  - GET    /documents              — список всіх документів з пагінацією та фільтрацією
  - POST   /documents/batch-confirm — пакетне підтвердження документів
  - POST   /documents/{id}/copy    — копіювання документа
  - DELETE /documents/{id}         — видалити документ (тільки чернетку)
  - GET    /documents/export       — експорт документів (Excel/CSV)
  - GET    /documents/{id}/print   — дані для друку документа
"""

import io
import csv
from datetime import datetime
from decimal import Decimal
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query, status, Request
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from pydantic import BaseModel, Field
from sqlalchemy import select, desc, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from app.database import get_session
from app.models.invoice import Invoice, InvoiceItem, InvoiceStatus
from app.models.transfer import Transfer, TransferItem, TransferStatus
from app.models.write_off import WriteOff, WriteOffItem
from app.models.return_invoice import ReturnInvoice, ReturnInvoiceItem, ReturnInvoiceStatus
from app.models.purchase_order import PurchaseOrder, PurchaseOrderItem, PurchaseOrderStatus
from app.models.supplier import Supplier
from app.models.product import Product
from app.models.user import User
from app.models.inventory import Inventory, InventoryItem as InventoryItemModel
from app.services.auth_service import AuthService
from app.services.document_service import DocumentService, generate_invoice_number

router = APIRouter(
    prefix="/documents",
    tags=["Документи"],
)

# Схема Bearer токена для Swagger
security_scheme = HTTPBearer(auto_error=False)


# ── Pydantic моделі для запитів та відповідей ──────────────────────────────

class BatchConfirmRequest(BaseModel):
    """Запит на пакетне підтвердження документів."""
    document_type: str = Field(
        ...,
        description="Тип документа: invoice, transfer, write_off, return_invoice, purchase_order",
    )
    ids: list[str] = Field(
        ...,
        description="Список UUID документів для підтвердження",
    )


class BatchConfirmResponse(BaseModel):
    """Відповідь на пакетне підтвердження."""
    confirmed_count: int = Field(..., description="Кількість успішно підтверджених")
    errors: list[dict] = Field(default_factory=list, description="Помилки по кожному ID")


class DocumentPrintData(BaseModel):
    """Дані для друку документа."""
    header: dict = Field(..., description="Заголовок: номер, дата, постачальник тощо")
    items: list[dict] = Field(..., description="Таблиця товарів")
    footer: dict = Field(..., description="Підсумки: сума, кількість позицій")


# ── DEPENDENCY: Отримання користувача з підтримкою token query-параметра ──

async def get_current_user_optional(
    credentials: HTTPAuthorizationCredentials = Depends(security_scheme),
    token: Optional[str] = Query(None, description="JWT токен (для друку в новій вкладці)"),
    session: AsyncSession = Depends(get_session),
) -> User:
    """
    Dependency для отримання поточного користувача.
    Спочатку пробує взяти токен з Authorization header (Bearer).
    Якщо немає — пробує взяти з query-параметра `token`.
    Якщо і там немає — повертає 401.
    """
    token_str = None

    # Спроба 1: з Authorization header
    if credentials is not None:
        token_str = credentials.credentials

    # Спроба 2: з query-параметра token
    if token_str is None and token:
        token_str = token

    # Якщо токена немає зовсім
    if token_str is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Необхідна авторизація. Передайте токен через заголовок Authorization: Bearer <token> "
                   "або через query-параметр ?token=<token>",
        )

    # Декодуємо токен
    payload = AuthService.decode_access_token(token_str)
    user_id = payload.get("sub")

    if not user_id:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Недійсний токен: відсутній ідентифікатор користувача",
        )

    # Шукаємо користувача в БД
    from sqlalchemy import select as sql_select
    result = await session.execute(
        sql_select(User).where(User.id == user_id)
    )
    user = result.scalar_one_or_none()

    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Користувача не знайдено",
        )

    if not user.is_active:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Користувач деактивований",
        )

    return user


# ── Схеми для відповіді документа (щоб уникнути дублювання) ─────────────────

def _build_document_item(
    doc_id: UUID,
    doc_type: str,
    number: str,
    status: str,
    total_amount: float,
    supplier_name: str,
    supplier_id_str: str | None,
    created_at_str: str | None,
    purchase_total: float | None = None,
    created_by: str = "",
    created_by_name: str = "",
) -> dict:
    """Створює словник з даними документа для списку."""
    return {
        "id": str(doc_id),
        "document_type": doc_type,
        "document_number": number,
        "status": status,
        "total_amount": total_amount,
        "purchase_total": purchase_total,
        "supplier_name": supplier_name,
        "supplier_id": supplier_id_str,
        "created_at": created_at_str,
        "created_by": created_by,
        "created_by_name": created_by_name,
    }


# ── Допоміжні функції для фільтрації ────────────────────────────────────────

def _parse_iso_date(date_str: str | None) -> datetime | None:
    """
    Перетворює ISO-рядок дати в об'єкт datetime.
    Повертає None, якщо рядок порожній або None.
    """
    if not date_str:
        return None
    try:
        return datetime.fromisoformat(date_str)
    except (ValueError, TypeError):
        return None


def _parse_ids(ids_str: str | None) -> list[UUID] | None:
    """
    Парсить рядок з ID через кому в список UUID.
    Повертає None, якщо рядок порожній або None.
    """
    if not ids_str:
        return None
    uuid_list = []
    for part in ids_str.split(","):
        part = part.strip()
        if not part:
            continue
        try:
            uuid_list.append(UUID(part))
        except (ValueError, TypeError):
            # Ігноруємо невалідні UUID
            pass
    return uuid_list if uuid_list else None


def _apply_date_filter(query, model_field, date_from: datetime | None, date_to: datetime | None):
    """Застосовує фільтр за датою до SQLAlchemy запиту."""
    if date_from:
        query = query.where(model_field >= date_from)
    if date_to:
        query = query.where(model_field <= date_to)
    return query


def _apply_amount_filter(query, model_field, amount_from: float | None, amount_to: float | None):
    """Застосовує фільтр за сумою до SQLAlchemy запиту."""
    if amount_from is not None:
        query = query.where(model_field >= amount_from)
    if amount_to is not None:
        query = query.where(model_field <= amount_to)
    return query


def _apply_supplier_filter(query, model_field, supplier_id: str | None):
    """Застосовує фільтр за постачальником до SQLAlchemy запиту."""
    if supplier_id:
        try:
            sid = UUID(supplier_id)
            query = query.where(model_field == sid)
        except (ValueError, AttributeError):
            pass  # Ігноруємо невалідний UUID
    return query


# ── Генератори номерів для документів ──────────────────────────────────────

async def _generate_transfer_number(session: AsyncSession) -> str:
    """Генерує номер для переміщення. Формат: ПМ-{YYYYMMDD}-{XXX}."""
    today = datetime.utcnow().strftime("%Y%m%d")
    prefix = f"ПМ-{today}-"
    result = await session.execute(
        select(func.max(Transfer.number))
        .where(Transfer.number.like(f"{prefix}%"))
    )
    max_number = result.scalar()
    last_seq = int(max_number[-3:]) if max_number else 0
    return f"{prefix}{last_seq + 1:03d}"


async def _generate_write_off_number(session: AsyncSession) -> str:
    """Генерує номер для списання. Формат: СП-{YYYYMMDD}-{XXX}."""
    today = datetime.utcnow().strftime("%Y%m%d")
    prefix = f"СП-{today}-"
    result = await session.execute(
        select(func.max(WriteOff.number))
        .where(WriteOff.number.like(f"{prefix}%"))
    )
    max_number = result.scalar()
    last_seq = int(max_number[-3:]) if max_number else 0
    return f"{prefix}{last_seq + 1:03d}"


# ── ОСНОВНИЙ ЕНДПОІНТ: Список документів ──────────────────────────────────

@router.get("")
async def list_documents(
    # Пагінація
    page: int = Query(1, ge=1, description="Сторінка"),
    size: int = Query(20, ge=1, le=100, description="Елементів на сторінці"),
    # Фільтри (старі, зворотна сумісність)
    status: Optional[str] = Query(None, description="Фільтр за статусом (draft, confirmed, cancelled)"),
    document_type: Optional[str] = Query(None, description="Фільтр за типом (invoice, transfer, write_off, return_invoice, purchase_order, inventory)"),
    search: Optional[str] = Query(None, description="Пошук за номером документа"),
    # Нові фільтри
    date_from: Optional[str] = Query(None, description="Дата від (ISO формат, напр. 2026-07-01)"),
    date_to: Optional[str] = Query(None, description="Дата до (ISO формат, напр. 2026-07-24)"),
    supplier_id: Optional[str] = Query(None, description="ID постачальника (UUID)"),
    amount_from: Optional[float] = Query(None, description="Сума від"),
    amount_to: Optional[float] = Query(None, description="Сума до"),
    # Залежності
    session: AsyncSession = Depends(get_session),
    current_user = Depends(AuthService.get_current_user),
):
    """
    Повертає об'єднаний список всіх документів (БЕЗ чеків продажу):
    - Прибуткові накладні (Invoice)
    - Переміщення (Transfer)
    - Списання (WriteOff)
    - Повернення постачальнику (ReturnInvoice)
    - Замовлення постачальнику (PurchaseOrder)
    - Інвентаризації (Inventory)

    Відсортовані за датою створення (від нових до старих).
    Підтримує фільтрацію за типом, статусом, пошук за номером,
    а також нові фільтри: date_from, date_to, supplier_id, amount_from, amount_to.
    """
    # Парсимо дати з ISO формату
    dt_from = _parse_iso_date(date_from)
    dt_to = _parse_iso_date(date_to)
    if dt_to:
        # Встановлюємо кінець дня, щоб включити всі документи за цю дату
        dt_to = dt_to.replace(hour=23, minute=59, second=59, microsecond=999999)

    offset = (page - 1) * size
    all_documents = []

    # ─── Прибуткові накладні ────────────────────────────────────────────────
    if not document_type or document_type == 'invoice':
        invoice_query = (
            select(Invoice)
            .options(
                selectinload(Invoice.items).selectinload(InvoiceItem.product),
                selectinload(Invoice.supplier),
                selectinload(Invoice.creator),
            )
        )
        invoice_query = _apply_date_filter(invoice_query, Invoice.created_at, dt_from, dt_to)
        invoice_query = _apply_amount_filter(invoice_query, Invoice.total_amount, amount_from, amount_to)
        invoice_query = _apply_supplier_filter(invoice_query, Invoice.supplier_id, supplier_id)
        invoice_query = invoice_query.order_by(desc(Invoice.created_at))

        invoices_result = await session.execute(invoice_query)
        invoices = invoices_result.scalars().all()

        for inv in invoices:
            inv_status = inv.status.value if hasattr(inv.status, 'value') else str(inv.status)
            if status and inv_status != status:
                continue
            if search and search.lower() not in (inv.number or '').lower():
                continue

            supplier_name = inv.supplier.name if inv.supplier else ""

            # Розрахунок закупівельної суми: sum(cost_price * quantity)
            purchase_total = 0.0
            if inv.items:
                purchase_total = sum(
                    float(item.cost_price or 0) * float(item.quantity or 0)
                    for item in inv.items
                )

            all_documents.append(_build_document_item(
                doc_id=inv.id,
                doc_type="invoice",
                number=inv.number,
                status=inv_status,
                total_amount=float(inv.total_amount) if inv.total_amount else 0,
                supplier_name=supplier_name,
                supplier_id_str=str(inv.supplier_id) if inv.supplier_id else None,
                created_at_str=inv.created_at.isoformat() if inv.created_at else None,
                purchase_total=purchase_total,
                created_by=str(inv.created_by_id) if inv.created_by_id else "",
                created_by_name=inv.creator.name if inv.creator else "",
            ))

    # ─── Переміщення ─────────────────────────────────────────────────────────
    if not document_type or document_type == 'transfer':
        transfer_query = select(Transfer).options(
            selectinload(Transfer.items),
            selectinload(Transfer.creator),
        )
        transfer_query = _apply_date_filter(transfer_query, Transfer.created_at, dt_from, dt_to)
        transfer_query = transfer_query.order_by(desc(Transfer.created_at))

        transfers_result = await session.execute(transfer_query)
        transfers = transfers_result.scalars().all()

        for tr in transfers:
            tr_status = tr.status.value if hasattr(tr.status, 'value') else str(tr.status)
            if status and tr_status != status:
                continue
            if search and search.lower() not in (tr.number or '').lower():
                continue

            all_documents.append(_build_document_item(
                doc_id=tr.id,
                doc_type="transfer",
                number=tr.number,
                status=tr_status,
                total_amount=0,
                supplier_name="",
                supplier_id_str=None,
                created_at_str=tr.created_at.isoformat() if tr.created_at else None,
                created_by=str(tr.created_by_id) if tr.created_by_id else "",
                created_by_name=tr.creator.name if tr.creator else "",
            ))

    # ─── Списання ────────────────────────────────────────────────────────────
    if not document_type or document_type == 'write_off':
        wo_query = select(WriteOff).options(
            selectinload(WriteOff.items).selectinload(WriteOffItem.product),
            selectinload(WriteOff.creator),
        )
        wo_query = _apply_date_filter(wo_query, WriteOff.created_at, dt_from, dt_to)
        wo_query = _apply_amount_filter(wo_query, WriteOff.total_amount, amount_from, amount_to)
        wo_query = wo_query.order_by(desc(WriteOff.created_at))

        writeoffs_result = await session.execute(wo_query)
        writeoffs = writeoffs_result.scalars().all()

        for wo in writeoffs:
            if status and status != "confirmed":
                continue
            if search and search.lower() not in (wo.number or '').lower():
                continue

            total = 0.0
            if wo.items:
                for item in wo.items:
                    price = float(item.product.price) if item.product and item.product.price else 0
                    qty = float(item.quantity) if item.quantity else 0
                    total += price * qty

            all_documents.append(_build_document_item(
                doc_id=wo.id,
                doc_type="write_off",
                number=wo.number,
                status="confirmed",
                total_amount=float(wo.total_amount) if wo.total_amount else total,
                supplier_name="",
                supplier_id_str=None,
                created_at_str=wo.created_at.isoformat() if wo.created_at else None,
                created_by=str(wo.created_by_id) if wo.created_by_id else "",
                created_by_name=wo.creator.name if wo.creator else "",
            ))

    # ─── Повернення постачальнику ────────────────────────────────────────────
    if not document_type or document_type == 'return_invoice':
        return_query = select(ReturnInvoice).options(
            selectinload(ReturnInvoice.items),
            selectinload(ReturnInvoice.supplier),
            selectinload(ReturnInvoice.creator),
        )
        return_query = _apply_date_filter(return_query, ReturnInvoice.created_at, dt_from, dt_to)
        return_query = _apply_amount_filter(return_query, ReturnInvoice.total_amount, amount_from, amount_to)
        return_query = _apply_supplier_filter(return_query, ReturnInvoice.supplier_id, supplier_id)
        return_query = return_query.order_by(desc(ReturnInvoice.created_at))

        returns_result = await session.execute(return_query)
        returns = returns_result.scalars().all()

        for ri in returns:
            ri_status = ri.status.value if hasattr(ri.status, 'value') else str(ri.status)
            if status and ri_status != status:
                continue
            if search and search.lower() not in (ri.number or '').lower():
                continue

            supplier_name = ri.supplier.name if ri.supplier else ""
            all_documents.append(_build_document_item(
                doc_id=ri.id,
                doc_type="return_invoice",
                number=ri.number,
                status=ri_status,
                total_amount=float(ri.total_amount) if ri.total_amount else 0,
                supplier_name=supplier_name,
                supplier_id_str=str(ri.supplier_id) if ri.supplier_id else None,
                created_at_str=ri.created_at.isoformat() if ri.created_at else None,
                created_by=str(ri.created_by_id) if ri.created_by_id else "",
                created_by_name=ri.creator.name if ri.creator else "",
            ))

    # ─── Замовлення постачальнику ────────────────────────────────────────────
    if not document_type or document_type == 'purchase_order':
        order_query = select(PurchaseOrder).options(
            selectinload(PurchaseOrder.items),
            selectinload(PurchaseOrder.supplier),
            selectinload(PurchaseOrder.creator),
        )
        order_query = _apply_date_filter(order_query, PurchaseOrder.created_at, dt_from, dt_to)
        order_query = _apply_amount_filter(order_query, PurchaseOrder.total_amount, amount_from, amount_to)
        order_query = _apply_supplier_filter(order_query, PurchaseOrder.supplier_id, supplier_id)
        order_query = order_query.order_by(desc(PurchaseOrder.created_at))

        orders_result = await session.execute(order_query)
        orders = orders_result.scalars().all()

        for po in orders:
            po_status = po.status.value if hasattr(po.status, 'value') else str(po.status)
            if status and po_status != status:
                continue
            if search and search.lower() not in (po.number or '').lower():
                continue

            supplier_name = po.supplier.name if po.supplier else ""
            all_documents.append(_build_document_item(
                doc_id=po.id,
                doc_type="purchase_order",
                number=po.number,
                status=po_status,
                total_amount=float(po.total_amount) if po.total_amount else 0,
                supplier_name=supplier_name,
                supplier_id_str=str(po.supplier_id) if po.supplier_id else None,
                created_at_str=po.created_at.isoformat() if po.created_at else None,
                created_by=str(po.created_by_id) if po.created_by_id else "",
                created_by_name=po.creator.name if po.creator else "",
            ))

    # ─── Інвентаризації ──────────────────────────────────────────────────
    if not document_type or document_type == 'inventory':
        inv_query = (
            select(Inventory)
            .options(
                selectinload(Inventory.items),
                selectinload(Inventory.creator),
            )
        )
        inv_query = _apply_date_filter(inv_query, Inventory.created_at, dt_from, dt_to)
        inv_query = inv_query.order_by(desc(Inventory.created_at))

        inv_result = await session.execute(inv_query)
        inventories = inv_result.scalars().all()

        for inv in inventories:
            inv_status = inv.status.value if hasattr(inv.status, 'value') else str(inv.status)
            if status and inv_status != status:
                continue
            if search and search.lower() not in (inv.number or '').lower():
                continue
            # Розрахунок підсумків
            total_cost = sum(float(item.actual_quantity) * float(item.cost_price) for item in inv.items)
            total_selling = sum(float(item.actual_quantity) * float(item.price) for item in inv.items)
            deviation_total = sum(float(item.difference) * float(item.cost_price) for item in inv.items)

            all_documents.append({
                "id": str(inv.id),
                "document_type": "inventory",
                "document_number": inv.number,
                "status": inv_status,
                "total_amount": total_selling,
                "purchase_total": total_cost,
                "deviation_total": deviation_total,
                "supplier_name": inv.location or "",
                "supplier_id": None,
                "created_at": inv.created_at.isoformat() if inv.created_at else None,
                "created_by": str(inv.created_by_id) if inv.created_by_id else "",
                "created_by_name": inv.creator.name if inv.creator else "",
            })

    # Сортуємо за датою (від нових до старих)
    all_documents.sort(key=lambda d: d["created_at"] or "", reverse=True)

    total = len(all_documents)
    paginated = all_documents[offset:offset + size]

    return {
        "items": paginated,
        "total": total,
        "page": page,
        "size": size,
        "pages": max(1, (total + size - 1) // size),
    }


# ── 2. ПАКЕТНЕ ПІДТВЕРДЖЕННЯ ───────────────────────────────────────────────

@router.post("/batch-confirm")
async def batch_confirm_documents(
    data: BatchConfirmRequest,
    session: AsyncSession = Depends(get_session),
    current_user = Depends(AuthService.require_admin),
):
    """
    Пакетне підтвердження документів.

    Приймає тип документа (document_type) та список ID.
    Підтверджує всі документи зі списку, використовуючи існуючі сервіси.

    Підтримувані типи: invoice, transfer, write_off, return_invoice, purchase_order
    """
    doc_service = DocumentService(session)
    confirmed_count = 0
    errors = []

    for doc_id_str in data.ids:
        try:
            doc_id = UUID(doc_id_str)
        except (ValueError, TypeError):
            errors.append({
                "id": doc_id_str,
                "error": f"Некоректний UUID: '{doc_id_str}'",
            })
            continue

        try:
            if data.document_type == "invoice":
                await doc_service.confirm_invoice(doc_id)
            elif data.document_type == "transfer":
                await doc_service.confirm_transfer(doc_id)
            elif data.document_type == "write_off":
                await doc_service.confirm_write_off(doc_id)
            elif data.document_type == "return_invoice":
                await doc_service.confirm_return_invoice(doc_id)
            elif data.document_type == "purchase_order":
                # Для замовлень постачальнику — створюємо прибуткову накладну
                result = await session.execute(
                    select(PurchaseOrder)
                    .options(selectinload(PurchaseOrder.items))
                    .where(PurchaseOrder.id == doc_id)
                )
                order = result.scalar_one_or_none()
                if not order:
                    raise HTTPException(status_code=404, detail=f"Замовлення '{doc_id}' не знайдено")

                if order.status != PurchaseOrderStatus.DRAFT:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Замовлення '{order.number}' вже має статус '{order.status.value}'",
                    )

                # Створюємо прибуткову накладну
                from app.models.invoice import Invoice as InvoiceModel, PaymentMethod, InvoiceStatus as InvoiceStatusModel, InvoiceItem as InvoiceItemModel

                invoice_number = await generate_invoice_number(session)
                new_invoice = InvoiceModel(
                    number=invoice_number,
                    supplier_id=order.supplier_id,
                    invoice_date=order.order_date,
                    payment_method=PaymentMethod.CREDIT,
                    is_fiscal=order.is_fiscal,
                    notes=f"Автоматично створено із замовлення №{order.number}",
                    total_amount=order.total_amount,
                    status=InvoiceStatusModel.DRAFT,
                    created_by_id=current_user.id,
                )
                session.add(new_invoice)
                await session.flush()

                for item in order.items:
                    invoice_item = InvoiceItemModel(
                        invoice_id=new_invoice.id,
                        product_id=item.product_id,
                        quantity=item.quantity,
                        price=item.price,
                        total=item.total,
                    )
                    session.add(invoice_item)

                order.invoice_id = new_invoice.id
                order.status = PurchaseOrderStatus.CONFIRMED
            else:
                errors.append({
                    "id": doc_id_str,
                    "error": f"Невідомий тип документа: '{data.document_type}'",
                })
                continue

            confirmed_count += 1

        except HTTPException as e:
            errors.append({
                "id": doc_id_str,
                "error": e.detail,
            })
        except Exception as e:
            errors.append({
                "id": doc_id_str,
                "error": str(e),
            })

    await session.flush()

    return BatchConfirmResponse(
        confirmed_count=confirmed_count,
        errors=errors,
    )


# ── 3. ВИДАЛЕННЯ ДОКУМЕНТА ────────────────────────────────────────────────

@router.delete("/{document_id}", status_code=204)
async def delete_document(
    document_id: UUID,
    document_type: str = Query(
        ...,
        description="Тип документа: invoice, transfer, write_off, return_invoice, purchase_order",
    ),
    session: AsyncSession = Depends(get_session),
    current_user = Depends(AuthService.require_admin),
):
    """
    Видаляє документ (тільки чернетку).

    Визначає тип документа та викликає відповідний delete-ендпоінт.
    """
    if document_type == "invoice":
        from app.models.invoice import Invoice, InvoiceStatus as InvStatus
        result = await session.execute(
            select(Invoice).where(Invoice.id == document_id)
        )
        doc = result.scalar_one_or_none()
        if not doc:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Накладну з ID '{document_id}' не знайдено",
            )
        if doc.status != InvStatus.DRAFT:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Можна видалити тільки чернетку",
            )
        await session.delete(doc)

    elif document_type == "transfer":
        from app.models.transfer import Transfer, TransferStatus as TrStatus
        result = await session.execute(
            select(Transfer).where(Transfer.id == document_id)
        )
        doc = result.scalar_one_or_none()
        if not doc:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Переміщення з ID '{document_id}' не знайдено",
            )
        if doc.status != TrStatus.DRAFT:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Можна видалити тільки чернетку",
            )
        await session.delete(doc)

    elif document_type == "write_off":
        from app.models.write_off import WriteOff
        result = await session.execute(
            select(WriteOff).where(WriteOff.id == document_id)
        )
        doc = result.scalar_one_or_none()
        if not doc:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Списання з ID '{document_id}' не знайдено",
            )
        if getattr(doc, 'status', 'draft') != 'draft':
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Можна видалити тільки чернетку",
            )
        await session.delete(doc)

    elif document_type == "return_invoice":
        from app.models.return_invoice import ReturnInvoice, ReturnInvoiceStatus as RiStatus
        result = await session.execute(
            select(ReturnInvoice).where(ReturnInvoice.id == document_id)
        )
        doc = result.scalar_one_or_none()
        if not doc:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Повернення з ID '{document_id}' не знайдено",
            )
        if doc.status != RiStatus.DRAFT:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Можна видалити тільки чернетку",
            )
        await session.delete(doc)

    elif document_type == "purchase_order":
        from app.models.purchase_order import PurchaseOrder, PurchaseOrderStatus as PoStatus
        result = await session.execute(
            select(PurchaseOrder).where(PurchaseOrder.id == document_id)
        )
        doc = result.scalar_one_or_none()
        if not doc:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Замовлення з ID '{document_id}' не знайдено",
            )
        if doc.status != PoStatus.DRAFT:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Можна видалити тільки чернетку",
            )
        await session.delete(doc)

    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Невідомий тип документа: '{document_type}'. "
                   f"Доступні: invoice, transfer, write_off, return_invoice, purchase_order",
        )

    await session.flush()


# ── 3. КОПІЮВАННЯ ДОКУМЕНТА ────────────────────────────────────────────────

@router.post("/{document_id}/copy")
async def copy_document(
    document_id: UUID,
    document_type: str = Query(
        ...,
        description="Тип документа: invoice, transfer, write_off, return_invoice, purchase_order",
    ),
    session: AsyncSession = Depends(get_session),
    current_user = Depends(AuthService.require_admin),
):
    """
    Створює копію документа з новим номером та статусом 'draft'.

    Копіюються: всі позиції (items), постачальник (supplier), дата, нотатки.
    Новий номер генерується автоматично.
    """
    now = datetime.utcnow()

    if document_type == "invoice":
        result = await session.execute(
            select(Invoice)
            .options(
                selectinload(Invoice.items).selectinload(InvoiceItem.product),
                selectinload(Invoice.supplier),
            )
            .where(Invoice.id == document_id)
        )
        original = result.scalar_one_or_none()
        if not original:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Накладну з ID '{document_id}' не знайдено",
            )

        new_number = await generate_invoice_number(session)
        new_doc = Invoice(
            number=new_number,
            supplier_id=original.supplier_id,
            invoice_date=now,
            payment_method=original.payment_method,
            is_fiscal=original.is_fiscal,
            notes=f"Копія накладної №{original.number}. {original.notes or ''}",
            total_amount=original.total_amount,
            status=InvoiceStatus.DRAFT,
            created_by_id=current_user.id,
        )
        session.add(new_doc)
        await session.flush()

        for item in original.items:
            session.add(InvoiceItem(
                invoice_id=new_doc.id,
                product_id=item.product_id,
                quantity=item.quantity,
                price=item.price,
                total=item.total,
            ))

        await session.flush()
        result = await session.execute(
            select(Invoice)
            .options(selectinload(Invoice.items).selectinload(InvoiceItem.product), selectinload(Invoice.supplier))
            .where(Invoice.id == new_doc.id)
        )
        from app.schemas.invoice import InvoiceResponse
        return InvoiceResponse.model_validate(result.scalar_one())

    elif document_type == "transfer":
        result = await session.execute(
            select(Transfer)
            .options(selectinload(Transfer.items))
            .where(Transfer.id == document_id)
        )
        original = result.scalar_one_or_none()
        if not original:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Переміщення з ID '{document_id}' не знайдено",
            )

        new_number = await _generate_transfer_number(session)
        new_doc = Transfer(
            number=new_number,
            from_location=original.from_location,
            to_location=original.to_location,
            transfer_date=now,
            notes=f"Копія переміщення №{original.number}. {original.notes or ''}",
            status=TransferStatus.DRAFT,
            created_by_id=current_user.id,
        )
        session.add(new_doc)
        await session.flush()

        for item in original.items:
            session.add(TransferItem(
                transfer_id=new_doc.id,
                product_id=item.product_id,
                quantity=item.quantity,
            ))

        await session.flush()
        result = await session.execute(
            select(Transfer).options(selectinload(Transfer.items)).where(Transfer.id == new_doc.id)
        )
        from app.schemas.transfer import TransferResponse
        return TransferResponse.model_validate(result.scalar_one())

    elif document_type == "write_off":
        result = await session.execute(
            select(WriteOff)
            .options(selectinload(WriteOff.items).selectinload(WriteOffItem.product))
            .where(WriteOff.id == document_id)
        )
        original = result.scalar_one_or_none()
        if not original:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Списання з ID '{document_id}' не знайдено",
            )

        new_number = await _generate_write_off_number(session)
        new_doc = WriteOff(
            number=new_number,
            reason=original.reason,
            write_off_date=now,
            notes=f"Копія списання №{original.number}. {original.notes or ''}",
            status="draft",
            total_amount=original.total_amount,
            created_by_id=current_user.id,
        )
        session.add(new_doc)
        await session.flush()

        for item in original.items:
            session.add(WriteOffItem(
                write_off_id=new_doc.id,
                product_id=item.product_id,
                quantity=item.quantity,
            ))

        await session.flush()
        result = await session.execute(
            select(WriteOff).options(selectinload(WriteOff.items)).where(WriteOff.id == new_doc.id)
        )
        from app.schemas.write_off import WriteOffResponse
        return WriteOffResponse.model_validate(result.scalar_one())

    elif document_type == "return_invoice":
        result = await session.execute(
            select(ReturnInvoice)
            .options(
                selectinload(ReturnInvoice.items).selectinload(ReturnInvoiceItem.product),
                selectinload(ReturnInvoice.supplier),
            )
            .where(ReturnInvoice.id == document_id)
        )
        original = result.scalar_one_or_none()
        if not original:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Повернення з ID '{document_id}' не знайдено",
            )

        # Використовуємо існуючий генератор з return_invoices.py
        from app.api.v1.return_invoices import generate_return_number
        new_number = await generate_return_number(session)

        new_doc = ReturnInvoice(
            number=new_number,
            supplier_id=original.supplier_id,
            return_date=now,
            return_action=original.return_action,
            is_fiscal=original.is_fiscal,
            notes=f"Копія повернення №{original.number}. {original.notes or ''}",
            total_amount=original.total_amount,
            status=ReturnInvoiceStatus.DRAFT,
            created_by_id=current_user.id,
        )
        session.add(new_doc)
        await session.flush()

        for item in original.items:
            session.add(ReturnInvoiceItem(
                return_invoice_id=new_doc.id,
                product_id=item.product_id,
                quantity=item.quantity,
                price=item.price,
                total=item.total,
            ))

        await session.flush()
        result = await session.execute(
            select(ReturnInvoice)
            .options(selectinload(ReturnInvoice.items).selectinload(ReturnInvoiceItem.product), selectinload(ReturnInvoice.supplier))
            .where(ReturnInvoice.id == new_doc.id)
        )
        from app.schemas.return_invoice import ReturnInvoiceResponse
        return ReturnInvoiceResponse.model_validate(result.scalar_one())

    elif document_type == "purchase_order":
        result = await session.execute(
            select(PurchaseOrder)
            .options(
                selectinload(PurchaseOrder.items).selectinload(PurchaseOrderItem.product),
                selectinload(PurchaseOrder.supplier),
            )
            .where(PurchaseOrder.id == document_id)
        )
        original = result.scalar_one_or_none()
        if not original:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Замовлення з ID '{document_id}' не знайдено",
            )

        from app.api.v1.purchase_orders import generate_order_number
        new_number = await generate_order_number(session)

        new_doc = PurchaseOrder(
            number=new_number,
            supplier_id=original.supplier_id,
            order_date=now,
            expected_date=original.expected_date,
            is_fiscal=original.is_fiscal,
            notes=f"Копія замовлення №{original.number}. {original.notes or ''}",
            total_amount=original.total_amount,
            status=PurchaseOrderStatus.DRAFT,
            created_by_id=current_user.id,
        )
        session.add(new_doc)
        await session.flush()

        for item in original.items:
            session.add(PurchaseOrderItem(
                purchase_order_id=new_doc.id,
                product_id=item.product_id,
                quantity=item.quantity,
                price=item.price,
                total=item.total,
            ))

        await session.flush()
        result = await session.execute(
            select(PurchaseOrder)
            .options(selectinload(PurchaseOrder.items).selectinload(PurchaseOrderItem.product), selectinload(PurchaseOrder.supplier))
            .where(PurchaseOrder.id == new_doc.id)
        )
        from app.schemas.purchase_order import PurchaseOrderResponse
        return PurchaseOrderResponse.model_validate(result.scalar_one())

    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Невідомий тип документа: '{document_type}'. "
                   f"Доступні: invoice, transfer, write_off, return_invoice, purchase_order",
        )


# ── 4. ЕКСПОРТ ДОКУМЕНТІВ ──────────────────────────────────────────────────

async def _get_documents_detailed(
    session: AsyncSession,
    ids: list[UUID] | None = None,
    dt_from: datetime | None = None,
    dt_to: datetime | None = None,
) -> list[dict]:
    """
    Отримує деталізовані дані документів з усіма товарами.

    Кожен рядок — один товар з документа.
    Якщо `ids` передано — тільки вказані документи.
    Інакше — всі документи (з фільтром за датою).
    """
    rows = []
    use_ids_filter = ids is not None and len(ids) > 0

    # ─── Прибуткові накладні ────────────────────────────────────────────────
    inv_query = select(Invoice).options(
        selectinload(Invoice.items).selectinload(InvoiceItem.product),
        selectinload(Invoice.supplier),
    )
    if use_ids_filter:
        inv_query = inv_query.where(Invoice.id.in_(ids))
    inv_query = _apply_date_filter(inv_query, Invoice.created_at, dt_from, dt_to)
    inv_query = inv_query.order_by(desc(Invoice.created_at))

    for inv in (await session.execute(inv_query)).scalars().all():
        for item in inv.items:
            product = item.product
            rows.append({
                "Тип документа": "Прибуткова накладна",
                "Номер документа": inv.number,
                "Дата": inv.created_at.strftime("%d.%m.%Y") if inv.created_at else "",
                "Постачальник": inv.supplier.name if inv.supplier else "",
                "Статус": inv.status.value if hasattr(inv.status, 'value') else str(inv.status),
                "Спосіб оплати": inv.payment_method.value if inv.payment_method else "не вказано",
                "Назва товару": product.title if product else "Невідомий товар",
                "Штрих-код": product.barcode if product else "",
                "Кількість": float(item.quantity) if item.quantity else 0,
                "Ціна": float(item.price) if item.price else 0,
                "Сума": float(item.total) if item.total else 0,
            })

    # ─── Переміщення ─────────────────────────────────────────────────────────
    tr_query = select(Transfer).options(
        selectinload(Transfer.items),
    )
    if use_ids_filter:
        tr_query = tr_query.where(Transfer.id.in_(ids))
    tr_query = _apply_date_filter(tr_query, Transfer.created_at, dt_from, dt_to)
    tr_query = tr_query.order_by(desc(Transfer.created_at))

    for tr in (await session.execute(tr_query)).scalars().all():
        for item in tr.items:
            # Отримуємо продукт окремим запитом (бо TransferItem не має relationship з Product)
            prod_result = await session.execute(
                select(Product).where(Product.id == item.product_id)
            )
            product = prod_result.scalar_one_or_none()
            rows.append({
                "Тип документа": "Переміщення",
                "Номер документа": tr.number,
                "Дата": tr.created_at.strftime("%d.%m.%Y") if tr.created_at else "",
                "Постачальник": f"{tr.from_location} → {tr.to_location}",
                "Статус": tr.status.value if hasattr(tr.status, 'value') else str(tr.status),
                "Спосіб оплати": "",
                "Назва товару": product.title if product else "Невідомий товар",
                "Штрих-код": product.barcode if product else "",
                "Кількість": float(item.quantity) if item.quantity else 0,
                "Ціна": 0,
                "Сума": 0,
            })

    # ─── Списання ────────────────────────────────────────────────────────────
    wo_query = select(WriteOff).options(
        selectinload(WriteOff.items).selectinload(WriteOffItem.product),
    )
    if use_ids_filter:
        wo_query = wo_query.where(WriteOff.id.in_(ids))
    wo_query = _apply_date_filter(wo_query, WriteOff.created_at, dt_from, dt_to)
    wo_query = wo_query.order_by(desc(WriteOff.created_at))

    for wo in (await session.execute(wo_query)).scalars().all():
        for item in wo.items:
            product = item.product
            price = float(product.price) if product and product.price else 0
            qty = float(item.quantity) if item.quantity else 0
            rows.append({
                "Тип документа": "Списання",
                "Номер документа": wo.number,
                "Дата": wo.created_at.strftime("%d.%m.%Y") if wo.created_at else "",
                "Постачальник": wo.reason.value if hasattr(wo.reason, 'value') else str(wo.reason),
                "Статус": "confirmed",
                "Спосіб оплати": "",
                "Назва товару": product.title if product else "Невідомий товар",
                "Штрих-код": product.barcode if product else "",
                "Кількість": qty,
                "Ціна": price,
                "Сума": price * qty,
            })

    # ─── Повернення постачальнику ────────────────────────────────────────────
    ri_query = select(ReturnInvoice).options(
        selectinload(ReturnInvoice.items).selectinload(ReturnInvoiceItem.product),
        selectinload(ReturnInvoice.supplier),
    )
    if use_ids_filter:
        ri_query = ri_query.where(ReturnInvoice.id.in_(ids))
    ri_query = _apply_date_filter(ri_query, ReturnInvoice.created_at, dt_from, dt_to)
    ri_query = ri_query.order_by(desc(ReturnInvoice.created_at))

    for ri in (await session.execute(ri_query)).scalars().all():
        for item in ri.items:
            product = item.product
            rows.append({
                "Тип документа": "Повернення постачальнику",
                "Номер документа": ri.number,
                "Дата": ri.created_at.strftime("%d.%m.%Y") if ri.created_at else "",
                "Постачальник": ri.supplier.name if ri.supplier else "",
                "Статус": ri.status.value if hasattr(ri.status, 'value') else str(ri.status),
                "Спосіб оплати": "",
                "Назва товару": product.title if product else "Невідомий товар",
                "Штрих-код": product.barcode if product else "",
                "Кількість": float(item.quantity) if item.quantity else 0,
                "Ціна": float(item.price) if item.price else 0,
                "Сума": float(item.total) if item.total else 0,
            })

    # ─── Замовлення постачальнику ────────────────────────────────────────────
    po_query = select(PurchaseOrder).options(
        selectinload(PurchaseOrder.items).selectinload(PurchaseOrderItem.product),
        selectinload(PurchaseOrder.supplier),
    )
    if use_ids_filter:
        po_query = po_query.where(PurchaseOrder.id.in_(ids))
    po_query = _apply_date_filter(po_query, PurchaseOrder.created_at, dt_from, dt_to)
    po_query = po_query.order_by(desc(PurchaseOrder.created_at))

    for po in (await session.execute(po_query)).scalars().all():
        for item in po.items:
            product = item.product
            rows.append({
                "Тип документа": "Замовлення постачальнику",
                "Номер документа": po.number,
                "Дата": po.created_at.strftime("%d.%m.%Y") if po.created_at else "",
                "Постачальник": po.supplier.name if po.supplier else "",
                "Статус": po.status.value if hasattr(po.status, 'value') else str(po.status),
                "Спосіб оплати": "",
                "Назва товару": product.title if product else "Невідомий товар",
                "Штрих-код": product.barcode if product else "",
                "Кількість": float(item.quantity) if item.quantity else 0,
                "Ціна": float(item.price) if item.price else 0,
                "Сума": float(item.total) if item.total else 0,
            })

    return rows


async def _get_all_documents_flat(
    session: AsyncSession,
    status: str | None,
    document_type: str | None,
    search: str | None,
    dt_from: datetime | None,
    dt_to: datetime | None,
    supplier_id: str | None,
    amount_from: float | None,
    amount_to: float | None,
    ids: list[UUID] | None = None,
) -> list[dict]:
    """
    Отримує всі документи без пагінації (для експорту).
    Повертає спрощений плоский список з українськими назвами колонок.

    Якщо передано `ids` — фільтрує ТІЛЬКИ за IDs та датою (інші фільтри ігноруються).
    """
    all_docs = []

    # Визначаємо, чи потрібно застосовувати IDs фільтр
    use_ids_filter = ids is not None and len(ids) > 0

    # Прибуткові накладні
    if not document_type or document_type == 'invoice':
        inv_query = select(Invoice).options(selectinload(Invoice.supplier))

        if use_ids_filter:
            # IDs фільтр — пріоритетний, беремо тільки вказані документи
            inv_query = inv_query.where(Invoice.id.in_(ids))
            inv_query = _apply_date_filter(inv_query, Invoice.created_at, dt_from, dt_to)
        else:
            # Звичайні фільтри
            inv_query = _apply_date_filter(inv_query, Invoice.created_at, dt_from, dt_to)
            inv_query = _apply_amount_filter(inv_query, Invoice.total_amount, amount_from, amount_to)
            inv_query = _apply_supplier_filter(inv_query, Invoice.supplier_id, supplier_id)

        inv_query = inv_query.order_by(desc(Invoice.created_at))

        result = await session.execute(inv_query)
        for inv in result.scalars().all():
            s = inv.status.value if hasattr(inv.status, 'value') else str(inv.status)
            if not use_ids_filter:
                if status and s != status:
                    continue
                if search and search.lower() not in (inv.number or '').lower():
                    continue
            all_docs.append({
                "Тип": "Прибуткова накладна",
                "Номер": inv.number,
                "Статус": s,
                "Дата": inv.created_at.strftime("%d.%m.%Y %H:%M") if inv.created_at else "",
                "Постачальник": inv.supplier.name if inv.supplier else "",
                "Сума": float(inv.total_amount) if inv.total_amount else 0,
            })

    # Переміщення
    if not document_type or document_type == 'transfer':
        tr_query = select(Transfer)

        if use_ids_filter:
            tr_query = tr_query.where(Transfer.id.in_(ids))
            tr_query = _apply_date_filter(tr_query, Transfer.created_at, dt_from, dt_to)
        else:
            tr_query = _apply_date_filter(tr_query, Transfer.created_at, dt_from, dt_to)

        tr_query = tr_query.order_by(desc(Transfer.created_at))

        result = await session.execute(tr_query)
        for tr in result.scalars().all():
            s = tr.status.value if hasattr(tr.status, 'value') else str(tr.status)
            if not use_ids_filter:
                if status and s != status:
                    continue
                if search and search.lower() not in (tr.number or '').lower():
                    continue
            all_docs.append({
                "Тип": "Переміщення",
                "Номер": tr.number,
                "Статус": s,
                "Дата": tr.created_at.strftime("%d.%m.%Y %H:%M") if tr.created_at else "",
                "Постачальник": f"{tr.from_location} → {tr.to_location}",
                "Сума": 0,
            })

    # Списання
    if not document_type or document_type == 'write_off':
        wo_query = select(WriteOff)

        if use_ids_filter:
            wo_query = wo_query.where(WriteOff.id.in_(ids))
            wo_query = _apply_date_filter(wo_query, WriteOff.created_at, dt_from, dt_to)
        else:
            wo_query = _apply_date_filter(wo_query, WriteOff.created_at, dt_from, dt_to)
            wo_query = _apply_amount_filter(wo_query, WriteOff.total_amount, amount_from, amount_to)

        wo_query = wo_query.order_by(desc(WriteOff.created_at))

        result = await session.execute(wo_query)
        for wo in result.scalars().all():
            if not use_ids_filter:
                if status and status != "confirmed":
                    continue
                if search and search.lower() not in (wo.number or '').lower():
                    continue
            reason_label = wo.reason.value if hasattr(wo.reason, 'value') else str(wo.reason)
            all_docs.append({
                "Тип": "Списання",
                "Номер": wo.number,
                "Статус": "confirmed",
                "Дата": wo.created_at.strftime("%d.%m.%Y %H:%M") if wo.created_at else "",
                "Постачальник": reason_label,
                "Сума": float(wo.total_amount) if wo.total_amount else 0,
            })

    # Повернення постачальнику
    if not document_type or document_type == 'return_invoice':
        ri_query = select(ReturnInvoice).options(selectinload(ReturnInvoice.supplier))

        if use_ids_filter:
            ri_query = ri_query.where(ReturnInvoice.id.in_(ids))
            ri_query = _apply_date_filter(ri_query, ReturnInvoice.created_at, dt_from, dt_to)
        else:
            ri_query = _apply_date_filter(ri_query, ReturnInvoice.created_at, dt_from, dt_to)
            ri_query = _apply_amount_filter(ri_query, ReturnInvoice.total_amount, amount_from, amount_to)
            ri_query = _apply_supplier_filter(ri_query, ReturnInvoice.supplier_id, supplier_id)

        ri_query = ri_query.order_by(desc(ReturnInvoice.created_at))

        result = await session.execute(ri_query)
        for ri in result.scalars().all():
            s = ri.status.value if hasattr(ri.status, 'value') else str(ri.status)
            if not use_ids_filter:
                if status and s != status:
                    continue
                if search and search.lower() not in (ri.number or '').lower():
                    continue
            all_docs.append({
                "Тип": "Повернення постачальнику",
                "Номер": ri.number,
                "Статус": s,
                "Дата": ri.created_at.strftime("%d.%m.%Y %H:%M") if ri.created_at else "",
                "Постачальник": ri.supplier.name if ri.supplier else "",
                "Сума": float(ri.total_amount) if ri.total_amount else 0,
            })

    # Замовлення постачальнику
    if not document_type or document_type == 'purchase_order':
        po_query = select(PurchaseOrder).options(selectinload(PurchaseOrder.supplier))

        if use_ids_filter:
            po_query = po_query.where(PurchaseOrder.id.in_(ids))
            po_query = _apply_date_filter(po_query, PurchaseOrder.created_at, dt_from, dt_to)
        else:
            po_query = _apply_date_filter(po_query, PurchaseOrder.created_at, dt_from, dt_to)
            po_query = _apply_amount_filter(po_query, PurchaseOrder.total_amount, amount_from, amount_to)
            po_query = _apply_supplier_filter(po_query, PurchaseOrder.supplier_id, supplier_id)

        po_query = po_query.order_by(desc(PurchaseOrder.created_at))

        result = await session.execute(po_query)
        for po in result.scalars().all():
            s = po.status.value if hasattr(po.status, 'value') else str(po.status)
            if not use_ids_filter:
                if status and s != status:
                    continue
                if search and search.lower() not in (po.number or '').lower():
                    continue
            all_docs.append({
                "Тип": "Замовлення постачальнику",
                "Номер": po.number,
                "Статус": s,
                "Дата": po.created_at.strftime("%d.%m.%Y %H:%M") if po.created_at else "",
                "Постачальник": po.supplier.name if po.supplier else "",
                "Сума": float(po.total_amount) if po.total_amount else 0,
            })

    # Сортуємо за датою (від нових до старих)
    all_docs.sort(key=lambda d: d.get("Дата", "") or "", reverse=True)
    return all_docs


def _generate_excel(headers: list[str], rows: list[list]) -> bytes:
    """Генерує Excel файл з даними."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Документи"

    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    cell_alignment = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    # Автоширина колонок
    for col_idx, header in enumerate(headers, 1):
        max_length = len(header)
        for row_idx in range(2, len(rows) + 2):
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _generate_csv(headers: list[str], rows: list[list]) -> bytes:
    """Генерує CSV файл з BOM для Excel сумісності."""
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return output.getvalue().encode('utf-8-sig')


@router.get("/export")
async def export_documents(
    # Фільтри (ті ж, що й в list_documents)
    ids: Optional[str] = Query(None, description="Список ID через кому, якщо потрібно експортувати конкретні документи"),
    status: Optional[str] = Query(None, description="Фільтр за статусом"),
    document_type: Optional[str] = Query(None, description="Фільтр за типом документа"),
    search: Optional[str] = Query(None, description="Пошук за номером"),
    date_from: Optional[str] = Query(None, description="Дата від (ISO формат)"),
    date_to: Optional[str] = Query(None, description="Дата до (ISO формат)"),
    supplier_id: Optional[str] = Query(None, description="ID постачальника"),
    amount_from: Optional[float] = Query(None, description="Сума від"),
    amount_to: Optional[float] = Query(None, description="Сума до"),
    format: str = Query("excel", description="Формат: excel або csv"),
    detailed: bool = Query(False, description="Деталізований експорт з усіма товарами (кожен рядок — один товар)"),
    # Залежності
    session: AsyncSession = Depends(get_session),
    current_user = Depends(AuthService.get_current_user),
):
    """
    Експортує список документів у форматі Excel або CSV.

    Застосовує всі фільтри (включно з новими date_from/date_to/supplier_id/amount_from/amount_to),
    але БЕЗ пагінації — повертає всі документи.

    Якщо передано `ids` — експортуються ТІЛЬКИ документи з вказаними ID
    (інші фільтри, окрім дати, ігноруються).

    Якщо `detailed=True` — експортує деталізовані дані з усіма товарами.
    Кожен рядок містить один товар з документа.
    Колонки: Тип документа, Номер документа, Дата, Постачальник, Статус,
    Спосіб оплати, Назва товару, Штрих-код, Кількість, Ціна, Сума.

    Приклад використання:
      GET /documents/export?detailed=true&format=excel
      GET /documents/export?detailed=true&ids=uuid1,uuid2&format=csv
    """
    dt_from = _parse_iso_date(date_from)
    dt_to = _parse_iso_date(date_to)
    if dt_to:
        dt_to = dt_to.replace(hour=23, minute=59, second=59, microsecond=999999)

    # Парсимо IDs (пріоритетний фільтр)
    uuid_list = _parse_ids(ids)

    if detailed:
        # Деталізований експорт — кожен рядок це один товар
        detailed_docs = await _get_documents_detailed(
            session=session,
            ids=uuid_list,
            dt_from=dt_from,
            dt_to=dt_to,
        )
        headers = [
            "Тип документа", "Номер документа", "Дата", "Постачальник",
            "Статус", "Спосіб оплати", "Назва товару", "Штрих-код",
            "Кількість", "Ціна", "Сума",
        ]
        rows = [
            [
                d["Тип документа"], d["Номер документа"], d["Дата"],
                d["Постачальник"], d["Статус"], d.get("Спосіб оплати", ""),
                d["Назва товару"], d["Штрих-код"],
                d["Кількість"], d["Ціна"], d["Сума"],
            ]
            for d in detailed_docs
        ]
    else:
        # Звичайний плоский експорт
        all_docs = await _get_all_documents_flat(
            session=session,
            status=status,
            document_type=document_type,
            search=search,
            dt_from=dt_from,
            dt_to=dt_to,
            supplier_id=supplier_id,
            amount_from=amount_from,
            amount_to=amount_to,
            ids=uuid_list,
        )

        headers = ["Тип", "Номер", "Статус", "Дата", "Постачальник", "Сума"]
        rows = [
            [d["Тип"], d["Номер"], d["Статус"], d["Дата"], d["Постачальник"], d["Сума"]]
            for d in all_docs
        ]

    if format == "csv":
        csv_bytes = _generate_csv(headers, rows)
        filename = f"documents_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return StreamingResponse(
            iter([csv_bytes]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    else:
        excel_bytes = _generate_excel(headers, rows)
        filename = f"documents_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            iter([excel_bytes]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )


# ── 5. ДАНІ ДЛЯ ДРУКУ ДОКУМЕНТА ────────────────────────────────────────────

@router.get("/{document_id}/print")
async def print_document(
    document_id: UUID,
    document_type: str = Query(
        ...,
        description="Тип документа: invoice, transfer, write_off, return_invoice, purchase_order",
    ),
    session: AsyncSession = Depends(get_session),
    current_user = Depends(get_current_user_optional),
):
    """
    Повертає структуровані дані для друку документа.

    Формат відповіді:
    - header: номер документа, дата, постачальник / склади / причина
    - items: таблиця товарів (назва, кількість, ціна, сума)
    - footer: загальна сума, кількість позицій

    Авторизація: підтримує токен через Authorization header та query-параметр `token`.
    """
    if document_type == "invoice":
        result = await session.execute(
            select(Invoice)
            .options(
                selectinload(Invoice.items).selectinload(InvoiceItem.product),
                selectinload(Invoice.supplier),
            )
            .where(Invoice.id == document_id)
        )
        doc = result.scalar_one_or_none()
        if not doc:
            raise HTTPException(status_code=404, detail=f"Накладну '{document_id}' не знайдено")

        items_data = []
        for item in doc.items:
            items_data.append({
                "product_name": item.product.title if item.product else "Невідомий товар",
                "barcode": item.product.barcode if item.product else "",
                "quantity": float(item.quantity),
                "price": float(item.price),
                "total": float(item.total),
            })

        payment_label = doc.payment_method.value if doc.payment_method else "не вказано"
        return DocumentPrintData(
            header={
                "document_type": "Прибуткова накладна",
                "document_number": doc.number,
                "date": doc.invoice_date.strftime("%d.%m.%Y") if doc.invoice_date else "",
                "supplier": doc.supplier.name if doc.supplier else "—",
                "payment_method": payment_label,
                "status": doc.status.value,
            },
            items=items_data,
            footer={
                "total_amount": float(doc.total_amount) if doc.total_amount else 0,
                "total_quantity": sum(it["quantity"] for it in items_data),
                "total_items": len(items_data),
            },
        )

    elif document_type == "transfer":
        result = await session.execute(
            select(Transfer)
            .options(selectinload(Transfer.items))
            .where(Transfer.id == document_id)
        )
        doc = result.scalar_one_or_none()
        if not doc:
            raise HTTPException(status_code=404, detail=f"Переміщення '{document_id}' не знайдено")

        items_data = []
        for item in doc.items:
            prod_result = await session.execute(
                select(Product).where(Product.id == item.product_id)
            )
            product = prod_result.scalar_one_or_none()
            items_data.append({
                "product_name": product.title if product else "Невідомий товар",
                "barcode": product.barcode if product else "",
                "quantity": float(item.quantity),
                "price": 0,
                "total": 0,
            })

        return DocumentPrintData(
            header={
                "document_type": "Переміщення товару",
                "document_number": doc.number,
                "date": doc.transfer_date.strftime("%d.%m.%Y") if doc.transfer_date else "",
                "from_location": doc.from_location,
                "to_location": doc.to_location,
                "status": doc.status.value,
            },
            items=items_data,
            footer={
                "total_amount": 0,
                "total_quantity": sum(it["quantity"] for it in items_data),
                "total_items": len(items_data),
            },
        )

    elif document_type == "write_off":
        result = await session.execute(
            select(WriteOff)
            .options(selectinload(WriteOff.items).selectinload(WriteOffItem.product))
            .where(WriteOff.id == document_id)
        )
        doc = result.scalar_one_or_none()
        if not doc:
            raise HTTPException(status_code=404, detail=f"Списання '{document_id}' не знайдено")

        reason_names = {
            "expired": "Закінчився термін придатності",
            "damaged": "Пошкодження / бій",
            "defect": "Брак / дефект",
            "theft": "Крадіжка",
            "inventory": "Інвентаризація (нестача)",
            "other": "Інше",
        }
        reason_label = doc.reason.value if hasattr(doc.reason, 'value') else str(doc.reason)

        items_data = []
        for item in doc.items:
            price = float(item.product.price) if item.product and item.product.price else 0
            total = price * float(item.quantity)
            items_data.append({
                "product_name": item.product.title if item.product else "Невідомий товар",
                "barcode": item.product.barcode if item.product else "",
                "quantity": float(item.quantity),
                "price": price,
                "total": total,
            })

        return DocumentPrintData(
            header={
                "document_type": "Списання товару",
                "document_number": doc.number,
                "date": doc.write_off_date.strftime("%d.%m.%Y") if doc.write_off_date else "",
                "reason": reason_names.get(reason_label, reason_label),
                "notes": doc.notes or "",
            },
            items=items_data,
            footer={
                "total_amount": sum(it["total"] for it in items_data),
                "total_quantity": sum(it["quantity"] for it in items_data),
                "total_items": len(items_data),
            },
        )

    elif document_type == "return_invoice":
        result = await session.execute(
            select(ReturnInvoice)
            .options(
                selectinload(ReturnInvoice.items).selectinload(ReturnInvoiceItem.product),
                selectinload(ReturnInvoice.supplier),
            )
            .where(ReturnInvoice.id == document_id)
        )
        doc = result.scalar_one_or_none()
        if not doc:
            raise HTTPException(status_code=404, detail=f"Повернення '{document_id}' не знайдено")

        action_names = {
            "deduct_from_debt": "Списання з боргу постачальника",
            "add_to_cash": "Зачислення в касу",
            "exchange": "Обмін на інший товар",
        }
        action_label = doc.return_action.value if hasattr(doc.return_action, 'value') else str(doc.return_action)

        items_data = []
        for item in doc.items:
            items_data.append({
                "product_name": item.product.title if item.product else "Невідомий товар",
                "barcode": item.product.barcode if item.product else "",
                "quantity": float(item.quantity),
                "price": float(item.price),
                "total": float(item.total),
            })

        return DocumentPrintData(
            header={
                "document_type": "Повернення постачальнику",
                "document_number": doc.number,
                "date": doc.return_date.strftime("%d.%m.%Y") if doc.return_date else "",
                "supplier": doc.supplier.name if doc.supplier else "—",
                "action": action_names.get(action_label, action_label),
                "status": doc.status.value,
            },
            items=items_data,
            footer={
                "total_amount": float(doc.total_amount) if doc.total_amount else 0,
                "total_quantity": sum(it["quantity"] for it in items_data),
                "total_items": len(items_data),
            },
        )

    elif document_type == "purchase_order":
        result = await session.execute(
            select(PurchaseOrder)
            .options(
                selectinload(PurchaseOrder.items).selectinload(PurchaseOrderItem.product),
                selectinload(PurchaseOrder.supplier),
            )
            .where(PurchaseOrder.id == document_id)
        )
        doc = result.scalar_one_or_none()
        if not doc:
            raise HTTPException(status_code=404, detail=f"Замовлення '{document_id}' не знайдено")

        items_data = []
        for item in doc.items:
            items_data.append({
                "product_name": item.product.title if item.product else "Невідомий товар",
                "barcode": item.product.barcode if item.product else "",
                "quantity": float(item.quantity),
                "price": float(item.price),
                "total": float(item.total),
            })

        expected = doc.expected_date.strftime("%d.%m.%Y") if doc.expected_date else "не вказано"
        return DocumentPrintData(
            header={
                "document_type": "Замовлення постачальнику",
                "document_number": doc.number,
                "date": doc.order_date.strftime("%d.%m.%Y") if doc.order_date else "",
                "supplier": doc.supplier.name if doc.supplier else "—",
                "expected_date": expected,
                "status": doc.status.value,
            },
            items=items_data,
            footer={
                "total_amount": float(doc.total_amount) if doc.total_amount else 0,
                "total_quantity": sum(it["quantity"] for it in items_data),
                "total_items": len(items_data),
            },
        )

    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Невідомий тип документа: '{document_type}'. "
                   f"Доступні: invoice, transfer, write_off, return_invoice, purchase_order",
        )
